# =========================================================
# ENQUIRIES MODULE
# Document Section 4.1 — Step 1
# =========================================================

from datetime import datetime

from flask import (
    Blueprint,
    flash,
    redirect,
    render_template,
    request,
    url_for,
    send_file,
)

from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

from flask_login import (
    current_user,
    login_required,
)

from app import db
from app.models import (
    Client,
    Enquiry,
    User,
    Agent
)

from app.sales_scope import (
    is_admin_user,
    is_sales_user,
    scope_enquiries,
    get_enquiry_or_404,
)


# =========================================================
# BLUEPRINT
# =========================================================

enquiries_bp = Blueprint(
    "enquiries",
    __name__,
    url_prefix="/enquiries"
)


# =========================================================
# AUTO ENQUIRY REFERENCE GENERATOR
# Example:
# ENQ-2026-000001
# ENQ-2026-000002
# =========================================================

def generate_enquiry_reference():

    current_year = datetime.now().year

    prefix = (
        f"ENQ-{current_year}-"
    )

    last_enquiry = (
        db.session.execute(
            db.select(Enquiry)
            .where(
                Enquiry.enquiry_reference.like(
                    f"{prefix}%"
                )
            )
            .order_by(
                Enquiry.id.desc()
            )
        )
        .scalars()
        .first()
    )

    if not last_enquiry:

        next_number = 1

    else:

        try:
            last_number = int(
                last_enquiry
                .enquiry_reference
                .split("-")[-1]
            )

            next_number = (
                last_number + 1
            )

        except (
            ValueError,
            IndexError
        ):
            next_number = 1

    return (
        f"{prefix}"
        f"{next_number:06d}"
    )


    # =========================================================
# CREATE NEW ENQUIRY
# URL: /enquiries/add
# GET  -> Form open chestundi
# POST -> Form data database lo save chestundi
# =========================================================

@enquiries_bp.route(
    "/add",
    methods=["GET", "POST"]
)
@login_required
def add_enquiry():

    # -----------------------------------------
    # DROPDOWN DATA
    # -----------------------------------------

    clients = (
        db.session.execute(
            db.select(Client)
            .where(
                Client.is_archived.is_(False)
            )
            .where(
                Client.assigned_to_id == current_user.id
                if is_sales_user()
                else True
            )
            .order_by(
                Client.company_name.asc()
            )
        )
        .scalars()
        .all()
    )

    users = (
        [current_user]
        if is_sales_user()
        else (
            db.session.execute(
                db.select(User)
                .where(User.is_active_user.is_(True))
                .order_by(User.full_name.asc())
            )
            .scalars()
            .all()
        )
    )

    # -----------------------------------------
    # FORM SUBMITTED
    # -----------------------------------------

    if request.method == "POST":

        client_id = request.form.get(
            "client_id",
            type=int
        )

        origin = request.form.get(
            "origin",
            ""
        ).strip()

        destination = request.form.get(
            "destination",
            ""
        ).strip()

        origin_port = request.form.get(
            "origin_port",
    ""
        ).strip()

        destination_port = request.form.get(
            "destination_port",
    ""
        ).strip()
        mode_of_shipment = request.form.get(
            "mode_of_shipment",
            ""
        ).strip()

        cargo_description = request.form.get(
            "cargo_description",
            ""
        ).strip()

        cargo_weight_volume = request.form.get(
            "cargo_weight_volume",
            ""
        ).strip()

        expected_timeline = request.form.get("expected_timeline", "").strip()
        incoterms = request.form.get("incoterms", "").strip()
        additional_instructions = request.form.get("additional_instructions", "").strip()
        sales_coordinator_id = request.form.get("sales_coordinator_id", type=int)

        handled_by_id = (
            current_user.id
            if is_sales_user()
            else request.form.get(
                "handled_by_id",
                type=int
            )
        )

        # -------------------------------------
        # REQUIRED FIELD VALIDATION
        # -------------------------------------

        if not all(
            [
                client_id,
                origin,
                destination,
                mode_of_shipment,
                cargo_description,
                handled_by_id,
            ]
        ):
            flash(
                "Please complete all required fields.",
                "danger"
            )

            return render_template(
                "enquiries/add.html",
                clients=clients,
                users=users
            )

        # -------------------------------------
        # SECURITY:
        # CLIENT MUST EXIST AND BE ACTIVE
        # -------------------------------------

        client = db.session.get(
            Client,
            client_id
        )

        if (
            not client
            or client.is_archived
            or (
                is_sales_user()
                and client.assigned_to_id != current_user.id
            )
        ):
            flash(
                "Please select a valid client.",
                "danger"
            )

            return render_template(
                "enquiries/add.html",
                clients=clients,
                users=users
            )

        # -------------------------------------
        # SECURITY:
        # HANDLED-BY USER MUST BE ACTIVE
        # -------------------------------------

        handled_by = db.session.get(
            User,
            handled_by_id
        )

        if (
            not handled_by
            or not handled_by.is_active_user
        ):
            flash(
                "Please select a valid staff owner.",
                "danger"
            )

            return render_template(
                "enquiries/add.html",
                clients=clients,
                users=users
            )

        # -------------------------------------
        # CREATE ENQUIRY OBJECT
        # -------------------------------------

        enquiry = Enquiry(
            enquiry_reference=(
                generate_enquiry_reference()
            ),
            client_id=client.id,
            origin=origin,
            destination=destination,

            origin_port=origin_port or None,
            destination_port=destination_port or None,

            mode_of_shipment=mode_of_shipment,
            cargo_description=cargo_description,
            cargo_weight_volume=(
                cargo_weight_volume or None
            ),
            handled_by_id=handled_by.id,
            status="open",
            created_by_id=current_user.id,
        )

        db.session.add(
            enquiry
        )

        try:
            db.session.commit()

        except Exception:
            db.session.rollback()

            flash(
                "Unable to create enquiry. Please try again.",
                "danger"
            )

            return render_template(
                "enquiries/add.html",
                clients=clients,
                users=users
            )

        flash(
            (
                f"Enquiry "
                f"{enquiry.enquiry_reference} "
                f"created successfully."
            ),
            "success"
        )

        return redirect(
            url_for(
                "enquiries.enquiry_list"
            )
        )

    # -----------------------------------------
    # GET REQUEST:
    # JUST OPEN THE FORM
    # -----------------------------------------

    return render_template(
        "enquiries/add.html",
        clients=clients,
        users=users
    )
    # =========================================================
# VIEW ENQUIRY DETAILS
# URL: /enquiries/<enquiry_id>
#
# Purpose:
# One enquiry record ni full details tho open chestundi.
# =========================================================

# =========================================================
# VIEW ENQUIRY DETAILS
# URL: /enquiries/<enquiry_id>
# =========================================================

@enquiries_bp.route("/<int:enquiry_id>")
@login_required
def view_enquiry(enquiry_id):
    enquiry = get_enquiry_or_404(enquiry_id)
    
    # ఇక్కడ ఆల్రెడీ కొటేషన్ ఉందో లేదో చెక్ చేస్తున్నాం
    from app.models import Quotation, Agent
    existing_quotation = db.session.execute(
        db.select(Quotation).where(Quotation.enquiry_id == enquiry.id)
    ).scalars().first()
    
    try:
        agents = Agent.query.order_by(Agent.country.asc(), Agent.name.asc()).all()
    except Exception:
        agents = []

    return render_template(
        "enquiries/view.html",
        enquiry=enquiry,
        agents=agents,
        existing_quotation=existing_quotation  # దీన్ని టెంప్లేట్‌కి పంపుతున్నాం
    )

# =========================================================
# UPDATE ENQUIRY STATUS
# URL: /enquiries/<enquiry_id>/status
#
# POST only:
# Detail page nunchi selected status ni save chestundi.
# =========================================================

@enquiries_bp.route(
    "/<int:enquiry_id>/status",
    methods=["POST"]
)
@login_required
def update_enquiry_status(enquiry_id):

    # Database nunchi enquiry record load
    enquiry = get_enquiry_or_404(enquiry_id)

    # Form nunchi selected status receive
    new_status = request.form.get(
        "status",
        ""
    ).strip()

    # Manam allow chese statuses matrame
    allowed_statuses = {
        "open",
        "in_progress",
        "quoted",
        "closed",
    }

    # Invalid value manually send chesina reject
    if new_status not in allowed_statuses:

        flash(
            "Invalid enquiry status.",
            "danger"
        )

        return redirect(
            url_for(
                "enquiries.view_enquiry",
                enquiry_id=enquiry.id
            )
        )

    # New status enquiry object lo set
    enquiry.status = new_status

    try:
        db.session.commit()

    except Exception:
        db.session.rollback()

        flash(
            "Unable to update enquiry status.",
            "danger"
        )

        return redirect(
            url_for(
                "enquiries.view_enquiry",
                enquiry_id=enquiry.id
            )
        )

    flash(
        "Enquiry status updated successfully.",
        "success"
    )

    return redirect(
        url_for(
            "enquiries.view_enquiry",
            enquiry_id=enquiry.id
        )
    )
    # =========================================================
# EDIT ENQUIRY
# URL: /enquiries/<enquiry_id>/edit
#
# GET:
# Existing enquiry data form lo chupistundi.
#
# POST:
# Updated enquiry data database lo save chestundi.
# =========================================================

@enquiries_bp.route(
    "/<int:enquiry_id>/edit",
    methods=["GET", "POST"]
)
@login_required
def edit_enquiry(enquiry_id):

    # -----------------------------------------
    # LOAD EXISTING ENQUIRY
    # -----------------------------------------

    enquiry = get_enquiry_or_404(enquiry_id)

    # -----------------------------------------
    # LOAD ACTIVE CLIENTS FOR DROPDOWN
    # -----------------------------------------

    clients = (
        db.session.execute(
            db.select(Client)
            .where(
                Client.is_archived.is_(False)
            )
            .order_by(
                Client.company_name.asc()
            )
        )
        .scalars()
        .all()
    )

    # -----------------------------------------
    # LOAD ACTIVE USERS FOR DROPDOWN
    # -----------------------------------------

    users = (
        db.session.execute(
            db.select(User)
            .where(
                User.is_active_user.is_(True)
            )
            .order_by(
                User.full_name.asc()
            )
        )
        .scalars()
        .all()
    )

    # -----------------------------------------
    # FORM SUBMITTED
    # -----------------------------------------

    if request.method == "POST":

        client_id = request.form.get(
            "client_id",
            type=int
        )

        origin = request.form.get(
            "origin",
            ""
        ).strip()

        destination = request.form.get(
            "destination",
            ""
        ).strip()

        origin_port = request.form.get(
            "origin_port",
    ""
        ).strip()

        destination_port = request.form.get(
            "destination_port",
    ""
        ).strip()

        mode_of_shipment = request.form.get(
            "mode_of_shipment",
            ""
        ).strip()

        cargo_description = request.form.get(
            "cargo_description",
            ""
        ).strip()

        cargo_weight_volume = request.form.get(
            "cargo_weight_volume",
            ""
        ).strip()

        handled_by_id = request.form.get(
            "handled_by_id",
            type=int
        )

        # -------------------------------------
        # REQUIRED FIELD VALIDATION
        # -------------------------------------

        if not all(
            [
                client_id,
                origin,
                destination,
                mode_of_shipment,
                cargo_description,
                handled_by_id,
            ]
        ):

            flash(
                "Please complete all required fields.",
                "danger"
            )

            return render_template(
                "enquiries/edit.html",
                enquiry=enquiry,
                clients=clients,
                users=users
            )

        # -------------------------------------
        # VALIDATE CLIENT
        # -------------------------------------

        client = db.session.get(
            Client,
            client_id
        )

        if (
            not client
            or client.is_archived
        ):

            flash(
                "Please select a valid client.",
                "danger"
            )

            return render_template(
                "enquiries/edit.html",
                enquiry=enquiry,
                clients=clients,
                users=users
            )

        # -------------------------------------
        # VALIDATE STAFF OWNER
        # -------------------------------------

        handled_by = db.session.get(
            User,
            handled_by_id
        )

        if (
            not handled_by
            or not handled_by.is_active_user
        ):

            flash(
                "Please select a valid staff owner.",
                "danger"
            )

            return render_template(
                "enquiries/edit.html",
                enquiry=enquiry,
                clients=clients,
                users=users
            )

        # -------------------------------------
        # UPDATE EXISTING RECORD
        #
        # Important:
        # New Enquiry create cheyyatledu.
        # Existing enquiry fields matrame
        # change chestunnam.
        # -------------------------------------

        enquiry.client_id = client.id

        enquiry.origin = origin

        enquiry.destination = destination

        enquiry.origin_port = origin_port or None
        enquiry.destination_port = destination_port or None

        enquiry.mode_of_shipment = (
            mode_of_shipment
        )

        enquiry.cargo_description = (
            cargo_description
        )

        enquiry.cargo_weight_volume = (
            cargo_weight_volume or None
        )

        enquiry.handled_by_id = (
            handled_by.id
        )

        # -------------------------------------
        # SAVE CHANGES
        # -------------------------------------

        try:
            db.session.commit()

        except Exception:
            db.session.rollback()

            flash(
                "Unable to update enquiry. Please try again.",
                "danger"
            )

            return render_template(
                "enquiries/edit.html",
                enquiry=enquiry,
                clients=clients,
                users=users
            )

        # -------------------------------------
        # SUCCESS
        # -------------------------------------

        flash(
            (
                f"Enquiry "
                f"{enquiry.enquiry_reference} "
                f"updated successfully."
            ),
            "success"
        )

        return redirect(
            url_for(
                "enquiries.view_enquiry",
                enquiry_id=enquiry.id
            )
        )

    # -----------------------------------------
    # GET REQUEST
    #
    # Existing data tho edit form open.
    # -----------------------------------------

    return render_template(
        "enquiries/edit.html",
        enquiry=enquiry,
        clients=clients,
        users=users
    )


import os
import pdfkit
from flask import render_template, make_response, url_for, current_app
from flask_login import login_required
from datetime import datetime, timezone
from app.models import Enquiry

@enquiries_bp.route('/enquiries/<int:enquiry_id>/download-pdf')
@login_required
def download_enquiry_pdf(enquiry_id):
    enquiry = Enquiry.query.get_or_404(enquiry_id)

    # Quotation మోడ్యూల్ లాగే లోగోల కోసం అబ్సల్యూట్ ఫైల్ పాత్స్ సెట్ చేయడం
    logo_path = os.path.join(current_app.root_path, 'static', 'images', 'logo.png')
    logo_url = f"file:///{logo_path.replace('\\', '/')}"

    logo2_path = os.path.join(current_app.root_path, 'static', 'images', 'logo2.png')
    logo2_url = f"file:///{logo2_path.replace('\\', '/')}"

    current_time = datetime.now(timezone.utc)

    # టెంప్లేట్‌కి డేటాను రెండర్ చేయడం
    html_content = render_template(
        "enquiries/pdf_template.html",
        enquiry=enquiry,
        logo_url=logo_url,
        logo2_url=logo2_url,
        now=current_time
    )

    # 4 వైపులా 10mm (1 cm) మార్జిన్‌లతో సింగిల్ పేజీ PDF ఆప్షన్స్
    options = {
        'page-size': 'A4',
        'margin-top': '10mm',
        'margin-right': '10mm',
        'margin-bottom': '10mm',
        'margin-left': '10mm',
        'enable-local-file-access': None,
        'encoding': 'UTF-8',
        'no-outline': None
    }

    # విండోస్ / లైనక్స్ సిస్టమ్ బట్టి pdfkit కాన్ఫిగరేషన్
    if os.name == "nt":
        config = pdfkit.configuration(
            wkhtmltopdf=r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe"
        )
        pdf = pdfkit.from_string(html_content, False, configuration=config, options=options)
    else:
        pdf = pdfkit.from_string(html_content, False, options=options)

    response = make_response(pdf)
    response.headers["Content-Type"] = "application/pdf"
    # ఒకవేళ డైరెక్ట్ బ్రౌజర్‌లో ఓపెన్ అవ్వాలంటే 'attachment' బదులుగా 'inline' వాడవచ్చు
    response.headers["Content-Disposition"] = f'attachment; filename=Enquiry_{enquiry.enquiry_reference}.pdf'

    return response



@enquiries_bp.route('/enquiry/delete/<int:id>')
@login_required
def delete_enquiry(id):
    try:
        enquiry = Enquiry.query.get_or_404(id)
        db.session.delete(enquiry)
        db.session.commit()
        flash('Enquiry deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting enquiry: {e}', 'danger')
        
    return redirect(url_for('enquiries.enquiry_list'))

import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from flask import send_file

# =========================================================
# ENQUIRY LIST
# =========================================================

@enquiries_bp.route("/")
@login_required
def enquiry_list():
    selected_search = request.args.get("search", "").strip()
    selected_status = request.args.get("status", "").strip()
    selected_mode = request.args.get("mode", "").strip()
    selected_client_id = request.args.get("client_id", "").strip()
    selected_handled_by_id = request.args.get("handled_by_id", "").strip()

    # బేస్ క్వెరీ (డ్యూప్లికేట్ జాయిన్ రాకుండా కేవలం అవసరమైనప్పుడు క్లయింట్ జాయిన్ అయ్యేలా)
    base_query = scope_enquiries(db.select(Enquiry))

    # కార్డ్స్ కోసం కౌంట్స్ లెక్కించడం (ఎటువంటి ఎర్రర్స్ రాకుండా సేఫ్ సబ్‌క్వెరీస్)
    all_stats_query = scope_enquiries(db.select(Enquiry))
    total_enquiries_count = db.session.execute(db.select(db.func.count()).select_from(all_stats_query.subquery())).scalar() or 0
    open_count = db.session.execute(db.select(db.func.count()).select_from(all_stats_query.where(Enquiry.status.notin_(["closed", "cancelled", "converted"])).subquery())).scalar() or 0
    in_progress_count = db.session.execute(db.select(db.func.count()).select_from(all_stats_query.where(Enquiry.status == "in_progress").subquery())).scalar() or 0
    quoted_count = db.session.execute(db.select(db.func.count()).select_from(all_stats_query.where(Enquiry.status == "quoted").subquery())).scalar() or 0
    converted_count = db.session.execute(db.select(db.func.count()).select_from(all_stats_query.where(Enquiry.status == "converted").subquery())).scalar() or 0

    query = base_query

    # సెర్చ్ లేదా క్లయింట్ ఫిల్టర్ ఉన్నప్పుడు మాత్రమే క్లయింట్‌ని జాయిన్ చేయడం (DuplicateAlias ఎర్రర్ రాకుండా)
    if selected_search or selected_client_id:
        query = query.join(Client, Enquiry.client_id == Client.id, isouter=True)

    if selected_search:
        query = query.where(
            (Enquiry.enquiry_reference.ilike(f"%{selected_search}%")) |
            (Enquiry.origin.ilike(f"%{selected_search}%")) |
            (Enquiry.destination.ilike(f"%{selected_search}%")) |
            (Client.company_name.ilike(f"%{selected_search}%"))
        )

    if selected_status == "dashboard_open":
        query = query.where(Enquiry.status.notin_(["closed", "cancelled", "converted"]))
    elif selected_status:
        query = query.where(Enquiry.status == selected_status)

    if selected_mode:
        query = query.where(Enquiry.mode_of_shipment == selected_mode)

    if selected_client_id:
        query = query.where(Enquiry.client_id == int(selected_client_id))

    if selected_handled_by_id:
        query = query.where(Enquiry.handled_by_id == int(selected_handled_by_id))

    enquiries = (
        db.session.execute(
            query.order_by(
                Enquiry.created_at.desc()
            )
        )
        .scalars()
        .all()
    )

    clients_list = db.session.execute(
        db.select(Client)
        .where(Client.is_archived.is_(False))
        .where(Client.assigned_to_id == current_user.id if is_sales_user() else True)
        .order_by(Client.company_name.asc())
    ).scalars().all()

    users_list = (
        [current_user]
        if is_sales_user()
        else db.session.execute(
            db.select(User)
            .where(User.is_active_user.is_(True))
            .order_by(User.full_name.asc())
        ).scalars().all()
    )

    return render_template(
        "enquiries/list.html",
        enquiries=enquiries,
        clients_list=clients_list,
        users_list=users_list,
        total_enquiries_count=total_enquiries_count,
        open_count=open_count,
        in_progress_count=in_progress_count,
        quoted_count=quoted_count,
        converted_count=converted_count,
        selected_search=selected_search,
        selected_status=selected_status,
        selected_mode=selected_mode,
        selected_client_id=selected_client_id,
        selected_handled_by_id=selected_handled_by_id
    )

# =========================================================
# EXPORT EXCEL 
# =========================================================

@enquiries_bp.route("/export-excel")
@login_required
def export_enquiries_excel():
    search = request.args.get("search", "").strip()
    status = request.args.get("status", "").strip()
    mode = request.args.get("mode", "").strip()
    client_id = request.args.get("client_id", "").strip()
    handled_by_id = request.args.get("handled_by_id", "").strip()

    query = scope_enquiries(db.select(Enquiry)).join(Client, Enquiry.client_id == Client.id)

    if search:
        query = query.where(
            (Enquiry.enquiry_reference.ilike(f"%{search}%")) |
            (Enquiry.origin.ilike(f"%{search}%")) |
            (Enquiry.destination.ilike(f"%{search}%")) |
            (Client.company_name.ilike(f"%{search}%"))
        )
        
    if status == "dashboard_open":
        query = query.where(Enquiry.status.notin_(["closed", "cancelled", "converted"]))
    elif status:
        query = query.where(Enquiry.status == status)
        
    if mode:
        query = query.where(Enquiry.mode_of_shipment == mode)
    if client_id:
        query = query.where(Enquiry.client_id == int(client_id))
    if handled_by_id:
        query = query.where(Enquiry.handled_by_id == int(handled_by_id))

    enquiries = db.session.execute(query.order_by(Enquiry.created_at.desc())).scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enquiries Report"

    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7F1D1D", end_color="7F1D1D", fill_type="solid")
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
    )

    headers = ["Reference", "Client", "Route", "Mode", "Handled By", "Date", "Status"]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin

    ws.row_dimensions[1].height = 24

    for row_idx, e in enumerate(enquiries, 2):
        client_name = e.client.company_name if e.client else "-"
        route = f"{e.origin} -> {e.destination}"
        mode = e.mode_of_shipment.replace("_", " ").title() if e.mode_of_shipment else "-"
        handled_by = e.handled_by.full_name if e.handled_by else "-"
        date_str = e.enquiry_date.strftime("%d %b %Y") if e.enquiry_date else "-"
        status_val = e.status.replace("_", " ").title() if e.status else "-"

        row_data = [e.enquiry_reference, client_name, route, mode, handled_by, date_str, status_val]
        ws.append(row_data)

        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Segoe UI", size=10)
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.row_dimensions[row_idx].height = 20

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Enquiries_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )