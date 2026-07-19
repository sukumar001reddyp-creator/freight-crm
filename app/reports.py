from datetime import datetime

from io import BytesIO

from sqlalchemy import func, extract

from openpyxl import Workbook
from openpyxl.styles import Font
from flask import send_file

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

from flask import send_file

from flask import (
    Blueprint,
    render_template,
    request,
)
from flask_login import (
    login_required,
    current_user,
)

from app import db
from app.models import (
    Client,
    Enquiry,
    Quotation,
    Shipment,
    SupportTicket,
    User,
)

reports_bp = Blueprint(
    "reports",
    __name__,
    url_prefix="/reports",
)


# ==========================================================
# SALES USER CHECK
# ==========================================================

def is_sales_user():

    return getattr(
        current_user,
        "role",
        "",
    ) in {
        "sales",
        "sales_executive",
    }


# ==========================================================
# REPORTS DASHBOARD
# ==========================================================

@reports_bp.route("/")
@login_required
def dashboard():

    # ------------------------------------------
    # FILTER VALUES
    # ------------------------------------------

    from_date = request.args.get("from_date", "").strip()

    to_date = request.args.get("to_date", "").strip()

    coordinator = request.args.get(
        "coordinator",
        type=int,
    )

    category = request.args.get(
        "category",
        "",
    ).strip()

    client_status = request.args.get(
        "client_status",
        "",
    ).strip()

    shipment_status = request.args.get(
        "shipment_status",
        "",
    ).strip()

    # ------------------------------------------
    # BASE QUERIES
    # ------------------------------------------

    client_query = Client.query.filter(
        Client.is_archived.is_(False)
    )

    enquiry_query = Enquiry.query

    quotation_query = Quotation.query

    shipment_query = Shipment.query

    support_query = SupportTicket.query

    # ------------------------------------------
    # SALES RESTRICTION
    # ------------------------------------------

    if is_sales_user():

        client_query = client_query.filter(
            Client.assigned_to_id == current_user.id
        )

        enquiry_query = enquiry_query.filter(
            Enquiry.handled_by_id == current_user.id
        )

        shipment_query = shipment_query.filter(
            Shipment.handled_by_id == current_user.id
        )

    # ------------------------------------------
    # DATE FILTER
    # ------------------------------------------

    if from_date:

        from_dt = datetime.strptime(
            from_date,
            "%Y-%m-%d",
        )

        client_query = client_query.filter(
            Client.created_at >= from_dt
        )

        enquiry_query = enquiry_query.filter(
            Enquiry.created_at >= from_dt
        )

        quotation_query = quotation_query.filter(
            Quotation.created_at >= from_dt
        )

        shipment_query = shipment_query.filter(
            Shipment.created_at >= from_dt
        )

    if to_date:

        to_dt = datetime.strptime(
            to_date,
            "%Y-%m-%d",
        )

        client_query = client_query.filter(
            Client.created_at <= to_dt
        )

        enquiry_query = enquiry_query.filter(
            Enquiry.created_at <= to_dt
        )

        quotation_query = quotation_query.filter(
            Quotation.created_at <= to_dt
        )

        shipment_query = shipment_query.filter(
            Shipment.created_at <= to_dt
        )

            # ------------------------------------------
    # COORDINATOR FILTER
    # ------------------------------------------

    if coordinator:

        client_query = client_query.filter(
            Client.assigned_to_id == coordinator
        )

        enquiry_query = enquiry_query.filter(
            Enquiry.handled_by_id == coordinator
        )

        shipment_query = shipment_query.filter(
            Shipment.handled_by_id == coordinator
        )

    # ------------------------------------------
    # CLIENT CATEGORY FILTER
    # ------------------------------------------

    if category:

        client_query = client_query.filter(
            Client.category == category
        )

    # ------------------------------------------
    # CLIENT STATUS FILTER
    # ------------------------------------------

    if client_status:

        client_query = client_query.filter(
            Client.status == client_status
        )

    # ------------------------------------------
    # SHIPMENT STATUS FILTER
    # ------------------------------------------

    if shipment_status:

        shipment_query = shipment_query.filter(
            Shipment.shipment_status == shipment_status
        )

    # ------------------------------------------
    # DASHBOARD STATISTICS
    # ------------------------------------------

    stats = {

        "clients":
            client_query.count(),

        "enquiries":
            enquiry_query.count(),

        "quotations":
            quotation_query.count(),

        "shipments":
            shipment_query.count(),

        "active_clients":
            client_query.filter(
                Client.status == "active"
            ).count(),

        "pending_enquiries":
            enquiry_query.filter(
                Enquiry.status == "open"
            ).count(),

        "approved_quotes":
            quotation_query.filter(
                Quotation.status == "approved"
            ).count(),

        "open_support":
            support_query.filter(
                SupportTicket.status != "closed"
            ).count(),

    }

    # ------------------------------------------
    # REPORT TABLE
    # ------------------------------------------

    shipments = (
        shipment_query
        .order_by(
            Shipment.created_at.desc()
        )
        .limit(50)
        .all()
    )

    # ------------------------------------------
    # DROPDOWNS
    # ------------------------------------------

    coordinators = (
        User.query
        .filter(
            User.is_active_user.is_(True)
        )
        .order_by(
            User.full_name.asc()
        )
        .all()
    )

    categories = (
        db.session.query(
            Client.category
        )
        .distinct()
        .all()
    )

    categories = [
        row[0]
        for row in categories
        if row[0]
    ]

    shipment_statuses = (
        db.session.query(
            Shipment.shipment_status
        )
        .distinct()
        .all()
    )

    shipment_statuses = [
        row[0]
        for row in shipment_statuses
        if row[0]
    ]

        # ------------------------------------------
    # MONTHLY ANALYTICS
    # ------------------------------------------

    monthly_clients = (
        db.session.query(
            extract("month", Client.created_at),
            func.count(Client.id),
        )
        .group_by(
            extract("month", Client.created_at)
        )
        .all()
    )

    monthly_enquiries = (
        db.session.query(
            extract("month", Enquiry.created_at),
            func.count(Enquiry.id),
        )
        .group_by(
            extract("month", Enquiry.created_at)
        )
        .all()
    )

    monthly_shipments = (
        db.session.query(
            extract("month", Shipment.created_at),
            func.count(Shipment.id),
        )
        .group_by(
            extract("month", Shipment.created_at)
        )
        .all()
    )

    monthly_quotations = (
        db.session.query(
            extract("month", Quotation.created_at),
            func.count(Quotation.id),
        )
        .group_by(
            extract("month", Quotation.created_at)
        )
        .all()
    )
        # ------------------------------------------
    # CONVERT MONTHLY DATA FOR CHART.JS
    # ------------------------------------------

    def build_month_array(rows):

        values = [0] * 12

        for month, total in rows:

            if month:

                values[int(month) - 1] = total

        return values

    monthly_clients = build_month_array(
        monthly_clients
    )

    monthly_enquiries = build_month_array(
        monthly_enquiries
    )

    monthly_shipments = build_month_array(
        monthly_shipments
    )

    monthly_quotations = build_month_array(
        monthly_quotations
    )
        # ------------------------------------------
    # CHART DATA
    # ------------------------------------------

    chart_labels = [
        "Clients",
        "Enquiries",
        "Quotations",
        "Shipments",
        "Active Clients",
        "Pending Enquiries",
        "Approved Quotes",
        "Support Tickets",
    ]

    chart_values = [
        stats["clients"],
        stats["enquiries"],
        stats["quotations"],
        stats["shipments"],
        stats["active_clients"],
        stats["pending_enquiries"],
        stats["approved_quotes"],
        stats["open_support"],
    ]

    # ------------------------------------------
    # RENDER PAGE
    # ------------------------------------------

    return render_template(
        "reports/index.html",

        stats=stats,

        shipments=shipments,

        coordinators=coordinators,

        categories=categories,

        shipment_statuses=shipment_statuses,

        chart_labels=chart_labels,

        chart_values=chart_values,

        monthly_clients=monthly_clients,

        monthly_enquiries=monthly_enquiries,

        monthly_shipments=monthly_shipments,

        monthly_quotations=monthly_quotations,

        filters={

            "from_date": from_date,

            "to_date": to_date,

            "coordinator": coordinator,

            "category": category,

            "client_status": client_status,

            "shipment_status": shipment_status,

                    },

    )

@reports_bp.route("/export/excel")
@login_required
def export_excel():

    wb = Workbook()

    # =====================================================
    # CLIENTS SHEET
    # =====================================================

    ws = wb.active
    ws.title = "Clients"

    headers = [

        "Client Code",
        "Company",
        "Contact Person",
        "Email",
        "Phone",
        "Category",
        "Status",
        "Assigned To",
        "Created"

    ]

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(
            row=1,
            column=col,
            value=header,
        )

        cell.font = Font(
            bold=True
        )

    clients = (
        Client.query
        .order_by(
            Client.created_at.desc()
        )
        .all()
    )

    row = 2

    for client in clients:

        ws.cell(
            row=row,
            column=1,
            value=client.client_reference,
        )

        ws.cell(
            row=row,
            column=2,
            value=client.company_name,
        )

        ws.cell(
            row=row,
            column=3,
            value=client.contact_person_name,
        )

        ws.cell(
            row=row,
            column=4,
            value=client.email,
        )

        ws.cell(
            row=row,
            column=5,
            value=client.primary_phone,
        )

        ws.cell(
            row=row,
            column=6,
            value=client.category,
        )

        ws.cell(
            row=row,
            column=7,
            value=client.status,
        )

        ws.cell(
            row=row,
            column=8,
            value=(
                client.assigned_to.full_name
                if client.assigned_to
                else ""
            ),
        )

        ws.cell(
            row=row,
            column=9,
            value=client.created_at.strftime(
                "%d-%m-%Y"
            ),
        )

        row += 1
            # =====================================================
    # ENQUIRIES SHEET
    # =====================================================

    ws = wb.create_sheet("Enquiries")

    headers = [

        "Reference",
        "Client",
        "Origin",
        "Destination",
        "Mode",
        "Cargo",
        "Status",
        "Handled By",
        "Created"

    ]

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(
            row=1,
            column=col,
            value=header,
        )

        cell.font = Font(bold=True)

    enquiries = (
        Enquiry.query
        .order_by(
            Enquiry.created_at.desc()
        )
        .all()
    )

    row = 2

    for enquiry in enquiries:

        ws.cell(row=row, column=1,
                value=enquiry.enquiry_reference)

        ws.cell(row=row, column=2,
                value=enquiry.client.company_name
                if enquiry.client else "")

        ws.cell(row=row, column=3,
                value=enquiry.origin)

        ws.cell(row=row, column=4,
                value=enquiry.destination)

        ws.cell(row=row, column=5,
                value=enquiry.mode_of_shipment)

        ws.cell(row=row, column=6,
                value=enquiry.cargo_description)

        ws.cell(row=row, column=7,
                value=enquiry.status)

        ws.cell(row=row, column=8,
                value=enquiry.handled_by.full_name
                if enquiry.handled_by else "")

        ws.cell(
            row=row,
            column=9,
            value=enquiry.created_at.strftime("%d-%m-%Y"),
        )

        row += 1


    # =====================================================
    # QUOTATIONS SHEET
    # =====================================================

    ws = wb.create_sheet("Quotations")

    headers = [

        "Quotation No",
        "Client",
        "Amount",
        "Currency",
        "Status",
        "Created"

    ]

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(
            row=1,
            column=col,
            value=header,
        )

        cell.font = Font(bold=True)

    quotations = (
        Quotation.query
        .order_by(
            Quotation.created_at.desc()
        )
        .all()
    )

    row = 2

    for quotation in quotations:

        ws.cell(
            row=row,
            column=1,
            value=quotation.quotation_number,
        )

        ws.cell(
            row=row,
            column=2,
            value=quotation.enquiry.client.company_name
            if quotation.enquiry
            and quotation.enquiry.client
            else "",
        )

        ws.cell(
            row=row,
            column=3,
            value=quotation.quotation_amount,
        )

        ws.cell(
            row=row,
            column=4,
            value=quotation.currency,
        )

        ws.cell(
            row=row,
            column=5,
            value=quotation.status,
        )

        ws.cell(
            row=row,
            column=6,
            value=quotation.created_at.strftime(
                "%d-%m-%Y"
            ),
        )

        row += 1

            # =====================================================
    # SHIPMENTS SHEET
    # =====================================================

    ws = wb.create_sheet("Shipments")

    headers = [

        "Shipment Ref",
        "Client",
        "Origin",
        "Destination",
        "Mode",
        "Status",
        "Handled By",
        "Created"

    ]

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(
            row=1,
            column=col,
            value=header,
        )

        cell.font = Font(bold=True)

    shipments = (
        Shipment.query
        .order_by(
            Shipment.created_at.desc()
        )
        .all()
    )

    row = 2

    for shipment in shipments:

        ws.cell(
            row=row,
            column=1,
            value=shipment.shipment_reference,
        )

        ws.cell(
            row=row,
            column=2,
            value=shipment.client.company_name
            if shipment.client else "",
        )

        ws.cell(
            row=row,
            column=3,
            value=shipment.origin,
        )

        ws.cell(
            row=row,
            column=4,
            value=shipment.destination,
        )

        ws.cell(
            row=row,
            column=5,
            value=shipment.mode_of_shipment,
        )

        ws.cell(
            row=row,
            column=6,
            value=shipment.shipment_status,
        )

        ws.cell(
            row=row,
            column=7,
            value=shipment.handled_by.full_name
            if shipment.handled_by else "",
        )

        ws.cell(
            row=row,
            column=8,
            value=shipment.created_at.strftime(
                "%d-%m-%Y"
            ),
        )

        row += 1


    # =====================================================
    # SUPPORT TICKETS SHEET
    # =====================================================

    ws = wb.create_sheet("Support Tickets")

    headers = [

        "Ticket ID",
        "Client",
        "Subject",
        "Status",
        "Created"

    ]

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(
            row=1,
            column=col,
            value=header,
        )

        cell.font = Font(bold=True)

    tickets = (
        SupportTicket.query
        .order_by(
            SupportTicket.created_at.desc()
        )
        .all()
    )

    row = 2

    for ticket in tickets:

        ws.cell(
            row=row,
            column=1,
            value=ticket.id,
        )

        ws.cell(
            row=row,
            column=2,
            value=ticket.client.company_name
            if ticket.client else "",
        )

        ws.cell(
            row=row,
            column=3,
            value=ticket.subject,
        )

        ws.cell(
            row=row,
            column=4,
            value=ticket.status,
        )

        ws.cell(
            row=row,
            column=5,
            value=ticket.created_at.strftime(
                "%d-%m-%Y"
            ),
        )

        row += 1


    # =====================================================
    # AUTO COLUMN WIDTH
    # =====================================================

    for sheet in wb.worksheets:

        for column_cells in sheet.columns:

            length = max(
                len(str(cell.value or ""))
                for cell in column_cells
            )

            sheet.column_dimensions[
                column_cells[0].column_letter
            ].width = min(
                length + 4,
                40,
            )


    # =====================================================
    # DOWNLOAD
    # =====================================================

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return send_file(

        output,

        as_attachment=True,

        download_name="FreightCRM_Report.xlsx",

        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

    )


from flask import render_template
from io import BytesIO

@reports_bp.route("/export/pdf")
@login_required
def export_pdf():
    from io import BytesIO
    from datetime import datetime

    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    )
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    elements = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Title'],
        fontSize=18,
        textColor=colors.HexColor('#7f1d1d'),
        alignment=TA_CENTER,
        spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#64748b'),
        alignment=TA_CENTER,
        spaceAfter=18
    )
    section_style = ParagraphStyle(
        'SectionStyle',
        parent=styles['Heading2'],
        fontSize=13,
        textColor=colors.HexColor('#7f1d1d'),
        spaceBefore=6,
        spaceAfter=10
    )

    # ====================== HEADER ======================
    elements.append(Paragraph("<b>FREIGHT CRM</b>", title_style))
    elements.append(Paragraph(
        "Professional Logistics & Freight Management System<br/>"
        f"Generated on: {datetime.now().strftime('%d %B %Y, %H:%M')}",
        subtitle_style
    ))

    # ====================== SUMMARY ======================
    elements.append(Paragraph("SUMMARY", section_style))

    summary_data = [
        ["Item", "Count"],
        ["Clients", Client.query.count()],
        ["Enquiries", Enquiry.query.count()],
        ["Quotations", Quotation.query.count()],
        ["Shipments", Shipment.query.count()],
        ["Support Tickets", SupportTicket.query.count()],
    ]

    summary_table = Table(summary_data, colWidths=[350, 120])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7f1d1d')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fef2f2')),
        ('GRID', (0, 0), (-1, -1), 0.8, colors.HexColor('#b91c1c')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    elements.append(PageBreak())

    # Helper function
    def create_table(title, headers, rows, col_widths):
        elements.append(Paragraph(title, section_style))
        data = [headers] + rows
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7f1d1d')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fef2f2')]),
            ('GRID', (0, 0), (-1, -1), 0.6, colors.HexColor('#b91c1c')),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 16))

    # ====================== CLIENTS ======================
    client_rows = []
    for c in Client.query.order_by(Client.created_at.desc()).all():
        client_rows.append([
            c.client_reference or "-",
            (c.company_name or "-")[:40],
            (c.contact_person_name or "-")[:20],
            c.status or "-"
        ])
    create_table(
        "CLIENTS",
        ["Code", "Company", "Contact", "Status"],
        client_rows,
        [95, 220, 120, 70]
    )
    elements.append(PageBreak())

    # ====================== ENQUIRIES ======================
    enquiry_rows = []
    for e in Enquiry.query.order_by(Enquiry.created_at.desc()).all():
        enquiry_rows.append([
            e.enquiry_reference or "-",
            (e.client.company_name if e.client else "-")[:30],
            e.origin or "-",
            e.destination or "-",
            e.status or "-"
        ])
    create_table(
        "ENQUIRIES",
        ["Reference", "Client", "Origin", "Destination", "Status"],
        enquiry_rows,
        [95, 160, 80, 80, 70]
    )
    elements.append(PageBreak())

    # ====================== QUOTATIONS ======================
    quotation_rows = []
    for q in Quotation.query.order_by(Quotation.created_at.desc()).all():
        client_name = "-"
        if q.enquiry and q.enquiry.client:
            client_name = q.enquiry.client.company_name
        quotation_rows.append([
            q.quotation_number or "-",
            (client_name or "-")[:35],
            q.status or "-",
            q.created_at.strftime("%d-%m-%Y") if q.created_at else "-"
        ])
    create_table(
        "QUOTATIONS",
        ["Quotation No", "Client", "Status", "Created"],
        quotation_rows,
        [110, 220, 80, 80]
    )
    elements.append(PageBreak())

    # ====================== SHIPMENTS ======================
    shipment_rows = []
    for s in Shipment.query.order_by(Shipment.created_at.desc()).all():
        shipment_rows.append([
            s.shipment_reference or "-",
            (s.client.company_name if s.client else "-")[:30],
            s.shipment_status or "-",
            s.created_at.strftime("%d-%m-%Y") if s.created_at else "-"
        ])
    create_table(
        "SHIPMENTS",
        ["Shipment Ref", "Client", "Status", "Created"],
        shipment_rows,
        [110, 220, 90, 80]
    )
    elements.append(PageBreak())

    # ====================== SUPPORT TICKETS ======================
    ticket_rows = []
    for t in SupportTicket.query.order_by(SupportTicket.created_at.desc()).all():
        ticket_rows.append([
            str(t.id),
            (t.client.company_name if t.client else "-")[:25],
            (t.subject or "-")[:30],
            t.status or "-"
        ])
    create_table(
        "SUPPORT TICKETS",
        ["Ticket ID", "Client", "Subject", "Status"],
        ticket_rows,
        [70, 150, 180, 90]
    )

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"FreightCRM_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
        mimetype="application/pdf"
    )
        # ==========================================
    # CLIENTS
    # ==========================================

    elements.append(
        Paragraph(
            "<b>Clients</b>",
            styles["Heading2"],
        )
    )

    client_data = [[

        "Code",
        "Company",
        "Contact",
        "Status"

    ]]

    clients = (
        Client.query
        .order_by(
            Client.created_at.desc()
        )
        .all()
    )

    for client in clients:

        client_data.append([

            client.client_reference,

            client.company_name,

            client.contact_person_name,

            client.status,

        ])

    client_table = Table(client_data)

    client_table.setStyle(

        TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#7f1d1d")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("GRID",(0,0),(-1,-1),0.5,colors.grey),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

        ])

    )

    elements.append(client_table)
    elements.append(PageBreak())

    # ==========================================
    # ENQUIRIES
    # ==========================================

    elements.append(
        Paragraph(
            "<b>Enquiries</b>",
            styles["Heading2"],
        )
    )

    enquiry_data = [[

        "Reference",
        "Client",
        "Origin",
        "Destination",
        "Status"

    ]]

    enquiries = (
        Enquiry.query
        .order_by(
            Enquiry.created_at.desc()
        )
        .all()
    )

    for enquiry in enquiries:

        enquiry_data.append([

            enquiry.enquiry_reference,

            enquiry.client.company_name
            if enquiry.client else "",

            enquiry.origin,

            enquiry.destination,

            enquiry.status,

        ])

    enquiry_table = Table(enquiry_data)

    enquiry_table.setStyle(

        TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#7f1d1d")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("GRID",(0,0),(-1,-1),0.5,colors.grey),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

        ])

    )

    elements.append(enquiry_table)
    elements.append(PageBreak())

        # ==========================================
    # QUOTATIONS
    # ==========================================

    elements.append(
        Paragraph(
            "<b>Quotations</b>",
            styles["Heading2"],
        )
    )

    quotation_data = [[
        "Quotation No",
        "Client",
        "Status",
        "Created",
    ]]

    quotations = (
        Quotation.query
        .order_by(
            Quotation.created_at.desc()
        )
        .all()
    )

    for quotation in quotations:

        quotation_data.append([

            quotation.quotation_number,

            (
                quotation.enquiry.client.company_name
                if quotation.enquiry
                and quotation.enquiry.client
                else ""
            ),

            quotation.status,

            quotation.created_at.strftime(
                "%d-%m-%Y"
            ),

        ])

    quotation_table = Table(quotation_data)

    quotation_table.setStyle(

        TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#7f1d1d")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("GRID",(0,0),(-1,-1),0.5,colors.grey),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

        ])

    )

    elements.append(quotation_table)
    elements.append(PageBreak())


    # ==========================================
    # SHIPMENTS
    # ==========================================

    elements.append(
        Paragraph(
            "<b>Shipments</b>",
            styles["Heading2"],
        )
    )

    shipment_data = [[

        "Shipment Ref",
        "Client",
        "Status",
        "Created",

    ]]

    shipments = (
        Shipment.query
        .order_by(
            Shipment.created_at.desc()
        )
        .all()
    )

    for shipment in shipments:

        shipment_data.append([

            shipment.shipment_reference,

            shipment.client.company_name
            if shipment.client else "",

            shipment.shipment_status,

            shipment.created_at.strftime(
                "%d-%m-%Y"
            ),

        ])

    shipment_table = Table(shipment_data)

    shipment_table.setStyle(

        TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#7f1d1d")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("GRID",(0,0),(-1,-1),0.5,colors.grey),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

        ])

    )

    elements.append(shipment_table)
    elements.append(PageBreak())


    # ==========================================
    # SUPPORT TICKETS
    # ==========================================

    elements.append(
        Paragraph(
            "<b>Support Tickets</b>",
            styles["Heading2"],
        )
    )

    ticket_data = [[

        "Ticket ID",
        "Client",
        "Subject",
        "Status",

    ]]

    tickets = (
        SupportTicket.query
        .order_by(
            SupportTicket.created_at.desc()
        )
        .all()
    )

    for ticket in tickets:

        ticket_data.append([

            str(ticket.id),

            ticket.client.company_name
            if ticket.client else "",

            ticket.subject,

            ticket.status,

        ])

    ticket_table = Table(ticket_data)

    ticket_table.setStyle(

        TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#7f1d1d")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("GRID",(0,0),(-1,-1),0.5,colors.grey),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

        ])

    )

    elements.append(ticket_table)


    # ==========================================
    # BUILD PDF
    # ==========================================

    doc.build(elements)

    buffer.seek(0)

    return send_file(

        buffer,

        as_attachment=True,

        download_name="FreightCRM_Report.pdf",

        mimetype="application/pdf",

    )

