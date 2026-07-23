import os
from io import BytesIO
import io
import re
import uuid
from datetime import datetime, timezone

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    current_app,
    abort,
    send_file,
    send_from_directory,
)

from flask_login import (
    login_required,
    current_user,
)

from werkzeug.utils import secure_filename

from sqlalchemy import or_
from openpyxl import Workbook, load_workbook

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)

from app import db

from app.models import (
    Client,
    ClientActivity,
    ClientAttachment,
    ClientAuditLog,
    ClientNote,
    ClientStatusHistory,
    ClientPipelineHistory,
    ClientTask,
    User,
    Enquiry,
    Quotation,
    Shipment,
    ClientPortalUser,
)


clients_bp = Blueprint(
    "clients",
    __name__,
    url_prefix="/clients"
)


# =========================================================
# DOCUMENT-MATCHED OPTIONS (SECTION 2.1 REPLACED COMPLIANT)
# =========================================================

CLIENT_CATEGORIES = [
    ("Consignee / Importer", "Consignee / Importer"),
    ("Shipper / Exporter", "Shipper / Exporter"),
    ("Freight Forwarding Agent (overseas partner / co-loader)", "Freight Forwarding Agent (overseas partner / co-loader)"),
    ("NVOCC", "NVOCC"),
    ("Customs Broker", "Customs Broker"),
    ("Trucking / Haulage Company", "Trucking / Haulage Company"),
    ("Shipping Line / Airline (Carrier)", "Shipping Line / Airline (Carrier)"),
    ("Warehouse / 3PL Provider", "Warehouse / 3PL Provider"),
    ("Direct Corporate Client", "Direct Corporate Client"),
    ("Government / Institutional Client", "Government / Institutional Client"),
]


CLIENT_STATUSES = [
    ("lead", "Lead / Prospect"),
    ("new", "New Client"),
    ("active", "Active / Existing Client"),
    ("key", "Key / Strategic Client"),
    ("at_risk", "At-Risk Client"),
    ("dormant", "Dormant / Inactive Client"),
    ("churned", "Churned / Lost Client"),
    ("reactivated", "Reactivated / Win-Back Client"),
    ("referral", "Referral Client"),
]


SERVICE_OPTIONS = [
    ("air_freight", "Air Freight"),
    ("sea_freight", "Sea Freight"),
    ("land_freight", "Land Freight"),
    ("customs_clearance", "Customs Clearance"),
    ("warehousing", "Warehousing"),
    ("project_cargo", "Project Cargo"),
]


LEAD_SOURCES = [
    ("referral", "Referral"),
    ("website", "Website"),
    ("cold_call", "Cold Call"),
    ("exhibition", "Exhibition"),
    ("tender_notice", "Tender Notice"),
    ("existing_network", "Existing Network"),
    ("other", "Other"),
]


PRIORITY_LEVELS = [
    ("high", "High"),
    ("medium", "Medium"),
    ("low", "Low"),
]


PIPELINE_STAGES = [
    ("prospect", "Prospect"),
    ("qualified", "Qualified"),
    ("needs_analysis", "Needs Analysis"),
    ("proposal", "Proposal / Quotation"),
    ("negotiation", "Negotiation"),
    ("won", "Won"),
    ("lost", "Lost"),
]


ALLOWED_EXTENSIONS = {
    "pdf",
    "jpg",
    "jpeg",
    "docx",
}


# =========================================================
# HELPERS
# =========================================================

def parse_date(value):
    if not value:
        return None

    try:
        return datetime.strptime(
            value,
            "%Y-%m-%d"
        ).date()

    except ValueError:
        return None


def allowed_file(filename):
    return (
        "." in filename
        and filename.rsplit(".", 1)[1].lower()
        in ALLOWED_EXTENSIONS
    )


def get_upload_folder():
    upload_folder = os.path.join(
        current_app.root_path,
        "static",
        "uploads",
        "clients"
    )

    os.makedirs(
        upload_folder,
        exist_ok=True
    )

    return upload_folder


def log_client_audit(
    client_id,
    action_type,
    action,
    title,
    description=None
):
    audit_log = ClientAuditLog(
        client_id=client_id,
        action_type=action_type,
        action=action,
        title=title,
        description=description,
        performed_by_id=current_user.id,
    )

    db.session.add(audit_log)
    return audit_log


def get_form_options():
    if is_sales_user():
        owners = [current_user]
    else:
        owners = (
            User.query
            .filter_by(is_active_user=True)
            .order_by(User.full_name.asc())
            .all()
        )

    return {
        "categories": CLIENT_CATEGORIES,
        "statuses": CLIENT_STATUSES,
        "services": SERVICE_OPTIONS,
        "lead_sources": LEAD_SOURCES,
        "priorities": PRIORITY_LEVELS,
        "owners": owners,
    }


def is_admin_user():
    return getattr(current_user, "role", None) == "admin"


def is_sales_user():
    return getattr(current_user, "role", None) in {
        "sales",
        "sales_executive",
    }


def can_access_client(client):
    if is_admin_user():
        return True

    if is_sales_user():
        return client.assigned_to_id == current_user.id

    return True


def get_accessible_client_or_404(client_id):
    client = db.get_or_404(Client, client_id)

    if not can_access_client(client):
        abort(404)

    return client


def export_safe_filename(value):
    value = re.sub(
        r"[^A-Za-z0-9._-]+",
        "_",
        (value or "client").strip()
    )
    value = value.strip("._")
    return value[:80] or "client"


def export_text(value):
    if value is None:
        return ""

    if isinstance(value, (list, tuple, set)):
        return ", ".join(
            export_text(item)
            for item in value
        )

    return str(value)


def export_date(value):
    if not value:
        return ""

    if isinstance(value, datetime):
        return value.strftime(
            "%Y-%m-%d %H:%M"
        )

    try:
        return value.strftime(
            "%Y-%m-%d"
        )
    except AttributeError:
        return export_text(value)


def export_label(value, options=None):
    if value is None:
        return ""

    if options:
        return dict(options).get(
            value,
            export_text(value)
            .replace("_", " ")
            .title()
        )

    return (
        export_text(value)
        .replace("_", " ")
        .title()
    )


def get_client_export_data(client):
    owner = db.session.get(User, client.assigned_to_id)
    creator = db.session.get(User, client.created_by_id)

    activities = db.session.execute(db.select(ClientActivity).where(ClientActivity.client_id == client.id).order_by(ClientActivity.activity_date.desc(), ClientActivity.id.desc())).scalars().all()
    notes = db.session.execute(db.select(ClientNote).where(ClientNote.client_id == client.id).order_by(ClientNote.created_at.desc(), ClientNote.id.desc())).scalars().all()
    tasks = db.session.execute(db.select(ClientTask).where(ClientTask.client_id == client.id).order_by(ClientTask.due_date.desc(), ClientTask.id.desc())).scalars().all()
    documents = db.session.execute(db.select(ClientAttachment).where(ClientAttachment.client_id == client.id).order_by(ClientAttachment.uploaded_at.desc(), ClientAttachment.id.desc())).scalars().all()
    status_history = db.session.execute(db.select(ClientStatusHistory).where(ClientStatusHistory.client_id == client.id).order_by(ClientStatusHistory.changed_at.desc(), ClientStatusHistory.id.desc())).scalars().all()
    pipeline_history = db.session.execute(db.select(ClientPipelineHistory).where(ClientPipelineHistory.client_id == client.id).order_by(ClientPipelineHistory.moved_at.desc(), ClientPipelineHistory.id.desc())).scalars().all()
    audit_logs = db.session.execute(db.select(ClientAuditLog).where(ClientAuditLog.client_id == client.id).order_by(ClientAuditLog.created_at.desc(), ClientAuditLog.id.desc())).scalars().all()
    enquiries = db.session.execute(db.select(Enquiry).where(Enquiry.client_id == client.id).order_by(Enquiry.id.desc())).scalars().all()
    quotations = db.session.execute(db.select(Quotation).join(Enquiry, Quotation.enquiry_id == Enquiry.id).where(Enquiry.client_id == client.id).order_by(Quotation.id.desc())).scalars().all()
    shipments = db.session.execute(db.select(Shipment).where(Shipment.client_id == client.id).order_by(Shipment.id.desc())).scalars().all()

    profile = [
        ("Client ID", client.id),
        ("Client Reference", client.client_reference),
        ("Company Name", client.company_name),
        ("Category", export_label(client.category, CLIENT_CATEGORIES)),
        ("Status", export_label(client.status, CLIENT_STATUSES)),
        ("Pipeline Stage", export_label(client.pipeline_stage, PIPELINE_STAGES)),
        ("Contact Person", client.contact_person_name),
        ("Designation", client.designation),
        ("Primary Phone", client.primary_phone),
        ("Secondary Phone", client.secondary_phone),
        ("Email", client.email),
        ("Website", client.website_url),
        ("Address Line 1", client.address_line_1),
        ("Address Line 2", client.address_line_2),
        ("Industry / Sector", client.industry_sector),
        ("Company Registration No.", client.company_registration_number),
        ("Tax / VAT Number", client.tax_vat_number),
        ("License Number", client.license_number),
        ("Payment Terms", client.payment_terms),
        ("Services Needed", [dict(SERVICE_OPTIONS).get(item, export_label(item)) for item in (client.services_needed or [])]),
        ("Assigned Owner", owner.full_name if owner else "Unassigned"),
        ("Lead Source", export_label(client.lead_source, LEAD_SOURCES)),
        ("Priority", export_label(client.priority_level, PRIORITY_LEVELS)),
        ("Last Contact Date", export_date(client.last_contact_date)),
        ("Next Follow-Up Date", export_date(client.next_follow_up_date)),
        ("Tags", client.tags or []),
        ("Profile Notes", client.notes),
        ("Created By", creator.full_name if creator else ""),
        ("Date Added", export_date(getattr(client, "date_added", None))),
        ("Archived", "Yes" if client.is_archived else "No"),
    ]

    return {
        "profile": profile,
        "activities": activities,
        "notes": notes,
        "tasks": tasks,
        "documents": documents,
        "status_history": status_history,
        "pipeline_history": pipeline_history,
        "audit_logs": audit_logs,
        "enquiries": enquiries,
        "quotations": quotations,
        "shipments": shipments,
    }


def pdf_paragraph(value, style):
    safe_value = (
        export_text(value)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br/>")
    )
    return Paragraph(safe_value or "—", style)


def pdf_table(rows, widths, body_style, header=True):
    prepared = []
    for row in rows:
        prepared.append([pdf_paragraph(value, body_style) for value in row])

    table = Table(prepared, colWidths=widths, repeatRows=1 if header else 0, hAlign="LEFT")
    commands = [
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]

    if header:
        commands.extend([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ])

    table.setStyle(TableStyle(commands))
    return table


def excel_prepare_sheet(ws, title, headers):
    ws.title = title
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.alignment = Alignment(vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def excel_finish_sheet(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = export_text(cell.value)
            if value:
                max_length = max(max_length, min(len(value), 60))
        ws.column_dimensions[letter].width = max(12, min(max_length + 2, 45))


# =========================================================
# CLIENT LIST
# =========================================================

@clients_bp.route("/")
@login_required
def client_list():
    search = request.args.get("search", "").strip()
    category = request.args.get("category", "").strip()
    status = request.args.get("status", "").strip()
    assigned_to = request.args.get("assigned_to", type=int)
    priority = request.args.get("priority", "").strip()
    pipeline_stage = request.args.get("pipeline_stage", "").strip()
    sort = request.args.get("sort", "newest").strip()
    page = request.args.get("page", 1, type=int)

    if page < 1:
        page = 1

    query = Client.query.filter_by(is_archived=False)

    if is_sales_user():
        query = query.filter(Client.assigned_to_id == current_user.id)

    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Client.client_reference.ilike(search_term),
                Client.company_name.ilike(search_term),
                Client.contact_person_name.ilike(search_term),
                Client.email.ilike(search_term),
                Client.primary_phone.ilike(search_term),
                Client.industry_sector.ilike(search_term),
            )
        )

    if category:
        query = query.filter(Client.category == category)
    if status:
        query = query.filter(Client.status == status)
    if assigned_to and not is_sales_user():
        query = query.filter(Client.assigned_to_id == assigned_to)
    if priority:
        query = query.filter(Client.priority_level == priority)
    if pipeline_stage:
        query = query.filter(Client.pipeline_stage == pipeline_stage)

    if sort == "oldest":
        query = query.order_by(Client.date_added.asc())
    elif sort == "name_az":
        query = query.order_by(Client.company_name.asc())
    elif sort == "name_za":
        query = query.order_by(Client.company_name.desc())
    elif sort == "follow_up":
        query = query.order_by(Client.next_follow_up_date.is_(None), Client.next_follow_up_date.asc(), Client.id.desc())
    elif sort == "last_activity":
        query = query.order_by(Client.last_contact_date.is_(None), Client.last_contact_date.desc(), Client.id.desc())
    else:
        query = query.order_by(Client.date_added.desc(), Client.id.desc())

    pagination = query.paginate(page=page, per_page=10, error_out=False)
    clients = pagination.items

    if is_sales_user():
        owners = [current_user]
    else:
        owners = User.query.filter_by(is_active_user=True).order_by(User.full_name.asc()).all()

    stats_query = Client.query.filter(Client.is_archived.is_(False))
    if is_sales_user():
        stats_query = stats_query.filter(Client.assigned_to_id == current_user.id)

    total_clients = stats_query.count()
    active_clients = stats_query.filter(Client.status.in_(["active", "key", "reactivated"])).count()
    lead_clients = stats_query.filter(Client.status.in_(["lead", "new"])).count()
    at_risk_clients = stats_query.filter(Client.status == "at_risk").count()

    return render_template(
        "clients/list.html",
        clients=clients,
        pagination=pagination,
        total_clients=total_clients,
        active_clients=active_clients,
        lead_clients=lead_clients,
        at_risk_clients=at_risk_clients,
        statuses=CLIENT_STATUSES,
        services=SERVICE_OPTIONS,
        categories=CLIENT_CATEGORIES,
        priorities=PRIORITY_LEVELS,
        owners=owners,
        pipeline_stages=PIPELINE_STAGES,
        selected_search=search,
        selected_category=category,
        selected_status=status,
        selected_assigned_to=assigned_to,
        selected_priority=priority,
        selected_pipeline_stage=pipeline_stage,
        selected_sort=sort,
    )


# =========================================================
# BULK REASSIGN CLIENT OWNER
# =========================================================

@clients_bp.route("/bulk/reassign-owner", methods=["POST"])
@login_required
def bulk_reassign_owner():
    if not is_admin_user():
        flash("Only Admin users can bulk reassign client owners.", "danger")
        return redirect(url_for("clients.client_list"))

    raw_client_ids = request.form.getlist("client_ids")
    new_owner_id = request.form.get("new_owner_id", type=int)

    if not raw_client_ids or not new_owner_id:
        flash("Select at least one client and a new owner.", "warning")
        return redirect(url_for("clients.client_list"))

    try:
        client_ids = sorted({int(cid) for cid in raw_client_ids if str(cid).isdigit()})
    except (TypeError, ValueError):
        client_ids = []

    new_owner = db.session.get(User, new_owner_id)
    if not new_owner or not new_owner.is_active_user:
        flash("Selected owner is invalid or inactive.", "danger")
        return redirect(url_for("clients.client_list"))

    clients = Client.query.filter(Client.id.in_(client_ids), Client.is_archived.is_(False)).all()
    changed_count = 0

    try:
        for client in clients:
            if client.assigned_to_id == new_owner.id:
                continue
            old_owner_name = client.assigned_to.full_name if client.assigned_to else "Unassigned"
            client.assigned_to_id = new_owner.id

            log_client_audit(
                client_id=client.id,
                action_type="owner_assignment",
                action="reassigned",
                title="Client owner reassigned",
                description=f"From: {old_owner_name}\nTo: {new_owner.full_name}\nMethod: Bulk reassignment"
            )
            changed_count += 1

        db.session.commit()
        flash(f"{changed_count} client(s) reassigned successfully.", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception(e)
        flash(str(e), "danger")

    return redirect(url_for("clients.client_list"))


# =========================================================
# BULK EXPORT SELECTED CLIENTS TO EXCEL
# =========================================================

@clients_bp.route("/bulk/export/excel", methods=["POST"])
@login_required
def bulk_export_clients_excel():
    if not is_admin_user():
        flash("Only Admin users can bulk export clients.", "danger")
        return redirect(url_for("clients.client_list"))

    raw_client_ids = request.form.getlist("client_ids")
    try:
        client_ids = sorted({int(cid) for cid in raw_client_ids if str(cid).isdigit()})
    except (TypeError, ValueError):
        client_ids = []

    clients = Client.query.filter(Client.id.in_(client_ids), Client.is_archived.is_(False)).order_by(Client.company_name.asc()).all()
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Selected Clients"

    headers = [
        "Client Reference", "Company Name", "Contact Person", "Email", "Primary Phone", "Secondary Phone",
        "Category", "Status", "Pipeline Stage", "Priority", "Owner", "Industry", "Registration No", "VAT No", "License No", "Payment Terms", "Next Follow-Up", "Last Contact"
    ]
    sheet.append(headers)

    for client in clients:
        sheet.append([
            client.client_reference or "", client.company_name or "", client.contact_person_name or "", client.email or "", client.primary_phone or "", client.secondary_phone or "",
            client.category or "", client.status or "", client.pipeline_stage or "", client.priority_level or "", client.assigned_to.full_name if client.assigned_to else "", client.industry_sector or "",
            client.company_registration_number or "", client.tax_vat_number or "", client.license_number or "", client.payment_terms or "",
            client.next_follow_up_date.strftime("%Y-%m-%d") if client.next_follow_up_date else "", client.last_contact_date.strftime("%Y-%m-%d") if client.last_contact_date else ""
        ])

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C62828")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"selected_clients_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# =========================================================
# BULK TAG SELECTED CLIENTS
# =========================================================

@clients_bp.route("/bulk/tag", methods=["POST"])
@login_required
def bulk_tag_clients():
    if not is_admin_user():
        flash("Only Admin users can bulk tag clients.", "danger")
        return redirect(url_for("clients.client_list"))

    raw_client_ids = request.form.getlist("client_ids")
    tag_value = (request.form.get("tag_value") or "").strip()
    tag_action = (request.form.get("tag_action") or "add").strip().lower()

    try:
        client_ids = sorted({int(cid) for cid in raw_client_ids if str(cid).isdigit()})
    except (TypeError, ValueError):
        client_ids = []

    tag_value = re.sub(r"\s+", " ", tag_value).strip()
    if not tag_value or len(tag_value) > 50:
        flash("Valid tag value under 50 characters required.", "warning")
        return redirect(url_for("clients.client_list"))

    selected_clients = Client.query.filter(Client.id.in_(client_ids), Client.is_archived.is_(False)).all()
    changed_count = 0

    try:
        for client in selected_clients:
            current_tags = list(client.tags or [])
            matching_tag = next((t for t in current_tags if str(t).strip().casefold() == tag_value.casefold()), None)

            if tag_action == "add" and matching_tag is None:
                current_tags.append(tag_value)
                client.tags = current_tags
                log_client_audit(client.id, "tags", "added", "Client tag added", f"Tag: {tag_value}\nMethod: Bulk tagging")
                changed_count += 1
            elif tag_action == "remove" and matching_tag is not None:
                current_tags.remove(matching_tag)
                client.tags = current_tags
                log_client_audit(client.id, "tags", "removed", "Client tag removed", f"Tag: {tag_value}\nMethod: Bulk tagging")
                changed_count += 1

        db.session.commit()
        flash(f"Tag operational updates applied across {changed_count} clients.", "success")
    except Exception:
        db.session.rollback()
        flash("Unable to complete bulk tagging.", "danger")

    return redirect(url_for("clients.client_list"))


# =========================================================
# DOWNLOAD CLIENT IMPORT TEMPLATE & EXCEL IMPORT (UPDATED)
# =========================================================

@clients_bp.route("/import/template")
@login_required
def download_client_import_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Clients Import"

    # Updated Headers with Section 2.2 Fields Included Mapped
    headers = [
        "Company Name", "Category", "Status", "Contact Person", "Designation", "Primary Phone", "Secondary Phone",
        "Email", "Website", "Address Line 1", "Address Line 2", "Industry", "Services Needed", "Lead Source", "Priority", 
        "Assigned Owner Email", "Last Contact Date", "Next Follow Up Date", "Tags", "Notes",
        "Secondary Contact Name", "Secondary Contact Phone", "Secondary Contact Email", "Company Registration Number", "Tax / VAT Number", "License Number", "Payment Terms"
    ]
    sheet.append(headers)

    sample = [
        "ABC Logistics Pvt Ltd", "Consignee / Importer", "active", "John David", "Logistics Manager", "9876543210", "",
        "john@abclogistics.com", "https://abclogistics.com", "Hyderabad", "", "Logistics", "air_freight,sea_freight", "website", "high",
        "admin@company.com", "2026-07-01", "2026-07-20", "VIP,Import", "Sample Client",
        "Backup Contact", "9111111111", "backup@abclogistics.com", "REG-12345", "VAT-99999", "LIC-88888", "Credit (30 days)"
    ]
    sheet.append(sample)

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C62828")

    for column in sheet.columns:
        max_length = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_length + 2, 18), 40)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="Client_Import_Template.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@clients_bp.route("/import", methods=["POST"])
@login_required
def import_clients_excel():
    file = request.files.get("excel_file")
    if not file:
        flash("Please select an Excel file.", "danger")
        return redirect(url_for("clients.client_list"))

    try:
        workbook = load_workbook(file)
        sheet = workbook.active
        imported = 0
        skipped = 0
        skip_duplicates = (request.form.get("ignore_duplicates") == "1")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            
            # Padded to 27 length to safely hold the new data parameters
            padded_row = list(row) + [None] * (27 - len(row))
            email = str(padded_row[7]).strip().lower() if padded_row[7] else ""

            if skip_duplicates and email:
                if Client.query.filter_by(email=email).first():
                    skipped += 1
                    continue

            # Re-building secondary contact dict from row index fields safely
            sec_contact = {
                "name": str(padded_row[20]).strip() if padded_row[20] else None,
                "phone": str(padded_row[21]).strip() if padded_row[21] else None,
                "email": str(padded_row[22]).strip().lower() if padded_row[22] else None
            }

            client = Client(
                company_name=str(padded_row[0]).strip(),
                category=padded_row[1] or "Consignee / Importer",
                status=padded_row[2] or "lead",
                contact_person_name=padded_row[3] or "Not Provided",
                designation=padded_row[4] or "",
                primary_phone=str(padded_row[5]).strip(),
                secondary_phone=str(padded_row[6]).strip() if padded_row[6] else None,
                email=email,
                website_url=padded_row[8] or "",
                address_line_1=padded_row[9] or "Not Available",
                address_line_2=padded_row[10] or "",
                industry_sector=padded_row[11] or "",
                services_needed=[s.strip() for s in str(padded_row[12]).split(",") if s.strip()] if padded_row[12] else [],
                assigned_to_id=current_user.id,
                lead_source=padded_row[13] or "",
                priority_level=padded_row[14] or "medium",
                notes=padded_row[19] or "",
                tags=[t.strip() for t in str(padded_row[18]).split(",") if t.strip()] if padded_row[18] else [],
                
                # EXTENDED NEW FIELDS APPLIED FROM EXCEL SHEET DATA DESIGNS
                secondary_contact_details=sec_contact,
                company_registration_number=str(padded_row[23]).strip() if padded_row[23] else None,
                tax_vat_number=str(padded_row[24]).strip() if padded_row[24] else None,
                license_number=str(padded_row[25]).strip() if padded_row[25] else None,
                payment_terms=str(padded_row[26]).strip() if padded_row[26] else None,
                created_by_id=current_user.id
            )
            db.session.add(client)
            db.session.flush()
            
            client.client_reference = f"CLT-{datetime.now(timezone.utc).year}-{client.id:06d}"
            db.session.add(ClientStatusHistory(client_id=client.id, old_status=None, new_status=client.status, changed_by_id=current_user.id, remarks="Client imported from compliant layout."))
            imported += 1
            
        db.session.commit()
        flash(f"Successfully processed {imported} clients from Excel. {skipped} duplicates skipped.", "success")
    except Exception as error:
        db.session.rollback()
        print("EXCEL IMPORT ERROR LOG:", repr(error))
        flash("Import payload execution dropped out.", "danger")
        
    return redirect(url_for("clients.client_list"))


# =========================================================
# ADD CLIENT (COMPLIANT ACTION ROUTE)
# =========================================================

@clients_bp.route("/add", methods=["GET", "POST"])
@login_required
def add_client():
    options = get_form_options()

    if request.method == "POST":
        company_name = request.form.get("company_name", "").strip()
        category = request.form.get("category", "").strip()
        status = request.form.get("status", "").strip()
        contact_person_name = request.form.get("contact_person_name", "").strip()
        primary_phone = request.form.get("primary_phone", "").strip()
        email = request.form.get("email", "").strip().lower()
        address_line_1 = request.form.get("address_line_1", "").strip()
        assigned_to_id = current_user.id if is_sales_user() else request.form.get("assigned_to_id", type=int)
        services_needed = request.form.getlist("services_needed")

        if not all([company_name, category, status, contact_person_name, primary_phone, email, address_line_1, assigned_to_id]):
            flash("Please complete all required fields.", "danger")
            return render_template("clients/add.html", **options)

        # Build Secondary Contact Dict (JSON schema compliant configuration data)
        sec_contact = {
            "name": request.form.get("secondary_contact_name", "").strip() or None,
            "phone": request.form.get("secondary_contact_phone", "").strip() or None,
            "email": request.form.get("secondary_contact_email", "").strip().lower() or None
        }

        client = Client(
            company_name=company_name,
            category=category,
            status=status,
            contact_person_name=contact_person_name,
            designation=request.form.get("designation", "").strip() or None,
            primary_phone=primary_phone,
            secondary_phone=request.form.get("secondary_phone", "").strip() or None,
            email=email,
            website_url=request.form.get("website_url", "").strip() or None,
            address_line_1=address_line_1,
            address_line_2=request.form.get("address_line_2", "").strip() or None,
            industry_sector=request.form.get("industry_sector", "").strip() or None,
            services_needed=services_needed,
            assigned_to_id=assigned_to_id,
            lead_source=request.form.get("lead_source", "").strip() or None,
            last_contact_date=parse_date(request.form.get("last_contact_date")),
            next_follow_up_date=parse_date(request.form.get("next_follow_up_date")),
            priority_level=request.form.get("priority_level", "").strip() or None,
            notes=request.form.get("notes", "").strip() or None,
            tags=[t.strip() for t in request.form.get("tags", "").split(",") if t.strip()],
            created_by_id=current_user.id,
            
            # SAVING NEW PARAMETERS SECTION 2.2 TRACKING DATA
            secondary_contact_details=sec_contact,
            company_registration_number=request.form.get("company_registration_number", "").strip() or None,
            tax_vat_number=request.form.get("tax_vat_number", "").strip() or None,
            license_number=request.form.get("license_number", "").strip() or None,
            payment_terms=request.form.get("payment_terms", "").strip() or None
        )

        db.session.add(client)
        db.session.flush()

        # Document Required Serial Mapping (CLT-2026-000001 constraint index auto generation)
        client.client_reference = f"CLT-{datetime.now(timezone.utc).year}-{client.id:06d}"

        db.session.add(ClientStatusHistory(client_id=client.id, old_status=None, new_status=status, changed_by_id=current_user.id, remarks="Initial client setup initialization."))
        
        # Safe preservation of physical attachments loop logic
        for f in request.files.getlist("attachments"):
            if f and f.filename and allowed_file(f.filename):
                orig = secure_filename(f.filename)
                stored = f"{uuid.uuid4().hex}.{orig.rsplit('.', 1)[1].lower()}"
                f.save(os.path.join(get_upload_folder(), stored))
                db.session.add(ClientAttachment(client_id=client.id, original_filename=orig, stored_filename=stored, file_path=f"uploads/clients/{stored}", file_type=stored.rsplit('.', 1)[1], uploaded_by_id=current_user.id))

        try:
            db.session.commit()
            flash(
                f"{client.company_name} saved with system designation ID {client.client_reference}.",
                "success"
            )
            return redirect(url_for("clients.client_list"))

        except Exception as e:
            db.session.rollback()

            import traceback
            traceback.print_exc()

            print("=" * 80)
            print("CLIENT SAVE ERROR")
            print(e)
            print("=" * 80)

            raise

    return render_template("clients/add.html", **options)


# =========================================================
# EDIT CLIENT (COMPLIANT ACTION ROUTE)
# =========================================================

@clients_bp.route("/<int:client_id>/edit", methods=["GET", "POST"])
@login_required
def edit_client(client_id):
    client = db.get_or_404(Client, client_id)
    options = get_form_options()

    if request.method == "POST":
        company_name = request.form.get("company_name", "").strip()
        category = request.form.get("category", "").strip()
        new_status = request.form.get("status", "").strip()
        contact_person_name = request.form.get("contact_person_name", "").strip()
        primary_phone = request.form.get("primary_phone", "").strip()
        email = request.form.get("email", "").strip().lower()
        address_line_1 = request.form.get("address_line_1", "").strip()
        assigned_to_id = request.form.get("assigned_to_id", type=int)
        services_needed = request.form.getlist("services_needed")

        if not all([company_name, category, new_status, contact_person_name, primary_phone, email, address_line_1, assigned_to_id]):
            flash("Please complete all required fields.", "danger")
            return render_template("clients/edit.html", client=client, **options)

        if client.status != new_status:
            db.session.add(ClientStatusHistory(client_id=client.id, old_status=client.status, new_status=new_status, changed_by_id=current_user.id, remarks="Updated from core editor profile setup."))

        client.company_name = company_name
        client.category = category
        client.status = new_status
        client.contact_person_name = contact_person_name
        client.designation = request.form.get("designation", "").strip() or None
        client.primary_phone = primary_phone
        client.secondary_phone = request.form.get("secondary_phone", "").strip() or None
        client.email = email
        client.website_url = request.form.get("website_url", "").strip() or None
        client.address_line_1 = address_line_1
        client.address_line_2 = request.form.get("address_line_2", "").strip() or None
        client.industry_sector = request.form.get("industry_sector", "").strip() or None
        client.services_needed = services_needed
        client.assigned_to_id = assigned_to_id
        client.lead_source = request.form.get("lead_source", "").strip() or None
        client.last_contact_date = parse_date(request.form.get("last_contact_date"))
        client.next_follow_up_date = parse_date(request.form.get("next_follow_up_date"))
        client.priority_level = request.form.get("priority_level", "").strip() or None
        client.notes = request.form.get("notes", "").strip() or None
        client.tags = [t.strip() for t in request.form.get("tags", "").split(",") if t.strip()]

        # RE-MAP NEW CORE PARAMETERS ON UPDATE ACTION
        client.secondary_contact_details = {
            "name": request.form.get("secondary_contact_name", "").strip() or None,
            "phone": request.form.get("secondary_contact_phone", "").strip() or None,
            "email": request.form.get("secondary_contact_email", "").strip().lower() or None
        }
        client.company_registration_number = request.form.get("company_registration_number", "").strip() or None
        client.tax_vat_number = request.form.get("tax_vat_number", "").strip() or None
        client.license_number = request.form.get("license_number", "").strip() or None
        client.payment_terms = request.form.get("payment_terms", "").strip() or None

        log_client_audit(client.id, "client_record", "updated", "Client record updated", "Core parameter details modified inside core model editor.")

        try:
            db.session.commit()
            flash(f"{client.company_name} updated successfully.", "success")
            return redirect(url_for("clients.view_client", client_id=client.id))
        except Exception:
            db.session.rollback()
            flash("Unable to update client configuration parameters.", "danger")

    return render_template("clients/edit.html", client=client, **options)


# =========================================================
# ADDITIONAL WORKFLOW MANAGEMENT ROUTES (PRESERVED SAFELY)
# =========================================================

@clients_bp.route("/<int:client_id>/convert-to-active", methods=["POST"])
@login_required
def convert_to_active(client_id):
    client = get_accessible_client_or_404(client_id)
    old_status = client.status
    client.status = "active"
    db.session.add(ClientStatusHistory(client_id=client.id, old_status=old_status, new_status="active", changed_by_id=current_user.id, remarks=request.form.get("remarks", "").strip() or "Lead conversion."))
    log_client_audit(client.id, "client_conversion", "converted_to_active", "Client converted to active lifecycle configuration.")
    db.session.commit()
    flash(f"{client.company_name} converted successfully.", "success")
    return redirect(url_for("clients.view_client", client_id=client.id))


@clients_bp.route("/<int:client_id>/reassign-owner", methods=["POST"])
@login_required
def reassign_owner(client_id):
    if not is_admin_user(): abort(403)
    client = get_accessible_client_or_404(client_id)
    new_owner_id = request.form.get("assigned_to_id", type=int)
    new_owner = db.session.get(User, new_owner_id)
    
    if new_owner and new_owner.is_active_user:
        old_name = client.assigned_to.full_name if client.assigned_to else "Unassigned"
        client.assigned_to_id = new_owner.id
        log_client_audit(client.id, "owner_assignment", "reassigned", "Client assigned layout modified", f"From: {old_name}\nTo: {new_owner.full_name}")
        db.session.commit()
        flash("Owner reconfiguration successful.", "success")
    return redirect(url_for("clients.view_client", client_id=client.id))


@clients_bp.route("/<int:client_id>/move-pipeline-stage", methods=["POST"])
@login_required
def move_pipeline_stage(client_id):
    client = get_accessible_client_or_404(client_id)
    new_stage = request.form.get("pipeline_stage", "").strip()
    if new_stage in [s[0] for s in PIPELINE_STAGES]:
        db.session.add(ClientPipelineHistory(client_id=client.id, old_stage=client.pipeline_stage, new_stage=new_stage, remarks=request.form.get("remarks", ""), moved_by_id=current_user.id))
        client.pipeline_stage = new_stage
        db.session.commit()
        flash("Workflow index shifted successfully.", "success")
    return redirect(url_for("clients.view_client", client_id=client.id))


@clients_bp.route("/<int:client_id>/change-status", methods=["POST"])
@login_required
def change_status(client_id):
    client = get_accessible_client_or_404(client_id)
    new_status = request.form.get("new_status", "").strip()
    if new_status in [s[0] for s in CLIENT_STATUSES] and new_status != client.status:
        db.session.add(ClientStatusHistory(client_id=client.id, old_status=client.status, new_status=new_status, changed_by_id=current_user.id, remarks=request.form.get("remarks", "")))
        client.status = new_status
        db.session.commit()
        flash("Lifecycle status index saved.", "success")
    return redirect(url_for("clients.view_client", client_id=client.id))


@clients_bp.route("/<int:client_id>/activities/add", methods=["POST"])
@login_required
def add_activity(client_id):
    client = get_accessible_client_or_404(client_id)
    activity = ClientActivity(client_id=client.id, activity_type=request.form.get("activity_type"), subject=request.form.get("subject"), description=request.form.get("description"), activity_date=datetime.now(timezone.utc), created_by_id=current_user.id)
    db.session.add(activity)
    client.last_contact_date = activity.activity_date.date()
    db.session.commit()
    return redirect(url_for("clients.view_client", client_id=client.id) + "#activities")


@clients_bp.route("/<int:client_id>/notes/add", methods=["POST"])
@login_required
def add_note(client_id):
    client = get_accessible_client_or_404(client_id)
    db.session.add(ClientNote(client_id=client.id, note_text=request.form.get("note_text"), created_by_id=current_user.id))
    db.session.commit()
    return redirect(url_for("clients.view_client", client_id=client.id) + "#notes")


@clients_bp.route("/<int:client_id>/notes/<int:note_id>/delete", methods=["POST"])
@login_required
def delete_note(client_id, note_id):
    note = db.get_or_404(ClientNote, note_id)
    if note.created_by_id == current_user.id:
        db.session.delete(note)
        db.session.commit()
    return redirect(url_for("clients.view_client", client_id=client_id) + "#notes")


@clients_bp.route("/<int:client_id>/tasks/add", methods=["POST"])
@login_required
def add_task(client_id):
    client = get_accessible_client_or_404(client_id)
    due_date = datetime.strptime(request.form.get("due_date"), "%Y-%m-%dT%H:%M")
    task = ClientTask(client_id=client.id, title=request.form.get("title"), description=request.form.get("description"), task_type=request.form.get("task_type"), due_date=due_date, assigned_to_id=request.form.get("assigned_to_id", type=int), priority=request.form.get("priority", "medium"), status="pending", created_by_id=current_user.id)
    db.session.add(task)
    client.next_follow_up_date = due_date.date()
    db.session.commit()
    return redirect(url_for("clients.view_client", client_id=client.id) + "#tasks")


@clients_bp.route("/<int:client_id>/tasks/<int:task_id>/status", methods=["POST"])
@login_required
def update_task_status(client_id, task_id):
    task = db.get_or_404(ClientTask, task_id)
    task.status = request.form.get("status")
    task.completed_at = datetime.now(timezone.utc) if task.status == "completed" else None
    db.session.commit()
    return redirect(url_for("clients.view_client", client_id=client_id) + "#tasks")


@clients_bp.route("/<int:client_id>/documents/upload", methods=["POST"])
@login_required
def upload_documents(client_id):
    client = get_accessible_client_or_404(client_id)
    for f in request.files.getlist("attachments"):
        if f and f.filename and allowed_file(f.filename):
            orig = secure_filename(f.filename)
            stored = f"{uuid.uuid4().hex}.{orig.rsplit('.', 1)[1].lower()}"
            f.save(os.path.join(get_upload_folder(), stored))
            db.session.add(ClientAttachment(client_id=client.id, original_filename=orig, stored_filename=stored, file_path=f"uploads/clients/{stored}", file_type=stored.rsplit('.', 1)[1], uploaded_by_id=current_user.id))
    db.session.commit()
    return redirect(url_for("clients.view_client", client_id=client.id) + "#documents")


@clients_bp.route("/<int:client_id>/documents/<int:attachment_id>/download")
@login_required
def download_document(client_id, attachment_id):
    attachment = db.get_or_404(ClientAttachment, attachment_id)
    return send_from_directory(get_upload_folder(), attachment.stored_filename, as_attachment=True, download_name=attachment.original_filename)


@clients_bp.route("/<int:client_id>/documents/<int:attachment_id>/delete", methods=["POST"])
@login_required
def delete_document(client_id, attachment_id):
    attachment = db.get_or_404(ClientAttachment, attachment_id)
    if attachment.uploaded_by_id == current_user.id:
        try: os.remove(os.path.join(current_app.root_path, "static", attachment.file_path))
        except OSError: pass
        db.session.delete(attachment)
        db.session.commit()
    return redirect(url_for("clients.view_client", client_id=client_id) + "#documents")


@clients_bp.route("/<int:client_id>/tasks/<int:task_id>/delete", methods=["POST"])
@login_required
def delete_task(client_id, task_id):
    client = get_accessible_client_or_404(client_id)
    task = db.get_or_404(ClientTask, task_id)
    db.session.delete(task)
    db.session.commit()
    return redirect(url_for("clients.view_client", client_id=client.id) + "#tasks")


# =========================================================
# EXPORT CLIENT RECORD - PDF (UPDATED WITH COMPLETE DETAILS)
# =========================================================

@clients_bp.route("/<int:client_id>/export/pdf")
@login_required
def export_client_pdf(client_id):
    client = get_accessible_client_or_404(client_id)
    data = get_client_export_data(client)

    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f"Client Record - {client.company_name}",
        author="Freight CRM",
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "ExportTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0F172A"),
        spaceAfter=8,
    )

    section_style = ParagraphStyle(
        "ExportSection",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#B91C1C"),
        spaceBefore=12,
        spaceAfter=6,
    )

    body_style = ParagraphStyle(
        "ExportBody",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        leading=11,
        textColor=colors.HexColor("#334155"),
    )

    small_style = ParagraphStyle(
        "ExportSmall",
        parent=body_style,
        fontSize=7,
        leading=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#64748B"),
    )

    story = [
        Paragraph("FREIGHT CRM - COMPLETE CLIENT RECORD", title_style),
        Paragraph(
            f"<b>{client.company_name}</b> &nbsp; | &nbsp; Reference: {client.client_reference or 'N/A'} &nbsp; | &nbsp; Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            small_style
        ),
        Spacer(1, 5 * mm),
        Paragraph("Client Profile Details", section_style),
    ]

    # Render All Profile Fields safely into Table Rows
    profile_rows = [["Field Parameter", "Configured Value Data"]]
    for label, value in data["profile"]:
        profile_rows.append([label, export_text(value)])

    story.append(pdf_table(profile_rows, [60 * mm, 195 * mm], body_style))

    # Render Associated Audit Log History summary mapping 
    if data["audit_logs"]:
        story.append(Paragraph("System Audit Trail Summary", section_style))
        audit_rows = [["Date", "Action Type", "Action performed", "Performed By"]]
        for audit in data["audit_logs"][:10]: # Limit to latest 10 for clean layout fit
            audit_rows.append([
                export_date(audit.created_at),
                export_label(audit.action_type),
                audit.title,
                audit.performed_by.full_name if audit.performed_by else "System"
            ])
        story.append(pdf_table(audit_rows, [40 * mm, 50 * mm, 115 * mm, 50 * mm], body_style))

    document.build(story)
    buffer.seek(0)

    filename = f"{export_safe_filename(client.company_name)}_complete_profile.pdf"
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
        max_age=0,
    )


# =========================================================
# EXPORT CLIENT RECORD - EXCEL (UPDATED WITH COMPLETE DETAILS)
# =========================================================

@clients_bp.route("/<int:client_id>/export/excel")
@login_required
def export_client_excel(client_id):
    client = get_accessible_client_or_404(client_id)
    data = get_client_export_data(client)

    workbook = Workbook()
    
    # Sheet 1: Master Profile Configuration Data
    ws = workbook.active
    excel_prepare_sheet(ws, "Client Profile", ["Field Parameter", "Configured Value Data"])

    for label, value in data["profile"]:
        ws.append([label, export_text(value)])
    excel_finish_sheet(ws)

    # Sheet 2: Audit Logs Tracking Data
    if data["audit_logs"]:
        ws_audit = workbook.create_sheet(title="Audit Trail History")
        excel_prepare_sheet(ws_audit, "Audit Trail History", ["Date Timestamp", "Action Type", "Action Log Title", "Description", "Performed By User"])
        
        for audit in data["audit_logs"]:
            ws_audit.append([
                export_date(audit.created_at),
                export_label(audit.action_type),
                audit.title,
                audit.description or "",
                audit.performed_by.full_name if audit.performed_by else "System"
            ])
        excel_finish_sheet(ws_audit)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"{export_safe_filename(client.company_name)}_complete_profile.xlsx"
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
        max_age=0,
    )


# =========================================================
# VIEW CLIENT, ARCHIVE, MERGE, PORTAL & OPERATIONAL ACTIONS
# =========================================================

@clients_bp.route("/<int:client_id>")
@login_required
def view_client(client_id):
    client = get_accessible_client_or_404(client_id)

    active_users = (
        User.query
        .filter_by(is_active_user=True)
        .order_by(User.full_name.asc())
        .all()
    )

    # Client Enquiries
    client_enquiries = (
        Enquiry.query
        .filter_by(client_id=client.id)
        .order_by(Enquiry.id.desc())
        .all()
    )

    # Client Quotations
    client_quotations = (
        db.session.query(Quotation)
        .join(Enquiry, Quotation.enquiry_id == Enquiry.id)
        .filter(Enquiry.client_id == client.id)
        .order_by(Quotation.id.desc())
        .all()
    )

    # Client Shipments
    client_shipments = (
        Shipment.query
        .filter_by(client_id=client.id)
        .order_by(Shipment.id.desc())
        .all()
    )

    merge_candidates = []
    if is_admin_user():
        merge_candidates = (
            Client.query
            .filter(Client.id != client.id, Client.is_archived.is_(False))
            .order_by(Client.company_name.asc())
            .all()
        )

    client_history = []

    # 1. Activities
    for activity in client.activities:
        client_history.append({
            "type": "activity",
            "date": activity.activity_date,
            "title": activity.subject,
            "description": activity.description,
            "actor": activity.created_by.full_name if activity.created_by else "Unknown User",
            "meta": activity.activity_type_label,
            "icon": {"call": "telephone-fill", "email": "envelope-fill", "meeting": "people-fill"}.get(activity.activity_type, "activity"),
        })

    # 2. Status History
    for history in client.status_history:
        client_history.append({
            "type": "status",
            "date": history.changed_at,
            "title": f"Status changed: {history.old_status or 'None'} → {history.new_status}",
            "description": history.remarks,
            "actor": history.changed_by.full_name if history.changed_by else "Unknown User",
            "meta": "Status Change",
            "icon": "arrow-left-right",
        })

    # 3. Pipeline Movements
    for movement in client.pipeline_history:
        client_history.append({
            "type": "pipeline",
            "date": movement.moved_at,
            "title": f"Pipeline moved: {movement.old_stage or 'None'} → {movement.new_stage}",
            "description": movement.remarks,
            "actor": movement.moved_by.full_name if movement.moved_by else "Unknown User",
            "meta": "Pipeline Movement",
            "icon": "diagram-3-fill",
        })

    # 4. Notes
    for note in client.client_notes:
        client_history.append({
            "type": "note",
            "date": note.created_at,
            "title": "Internal note added",
            "description": note.note_text,
            "actor": note.created_by.full_name if note.created_by else "Unknown User",
            "meta": "Note / Remark",
            "icon": "chat-left-text-fill",
        })

    # 5. Tasks
    for task in client.tasks:
        client_history.append({
            "type": "task",
            "date": task.created_at,
            "title": task.title,
            "description": task.description,
            "actor": task.created_by.full_name if task.created_by else "Unknown User",
            "meta": f"{task.task_type.title()} • {task.status.title()}",
            "icon": "calendar2-check-fill",
        })

    def history_sort_key(item):
        val = item.get("date")
        if val is None:
            return datetime.min.replace(tzinfo=timezone.utc)
        if isinstance(val, datetime):
            return val.replace(tzinfo=timezone.utc) if val.tzinfo is None else val.astimezone(timezone.utc)
        return datetime.combine(val, datetime.min.time(), tzinfo=timezone.utc)

    client_history.sort(key=history_sort_key, reverse=True)

    return render_template(
        "clients/view.html",
        client=client,
        client_history=client_history,
        pipeline_stages=PIPELINE_STAGES,
        merge_candidates=merge_candidates,
        client_enquiries=client_enquiries,
        client_quotations=client_quotations,
        client_shipments=client_shipments,
        statuses=CLIENT_STATUSES,
        services=dict(SERVICE_OPTIONS),
        categories=dict(CLIENT_CATEGORIES),
        lead_sources=dict(LEAD_SOURCES),
        task_assignees=active_users,
        owner_assignees=active_users,
    )


@clients_bp.route("/<int:client_id>/portal-account", methods=["POST"])
@login_required
def create_portal_account(client_id):
    client = get_accessible_client_or_404(client_id)
    account = ClientPortalUser(client_id=client.id, email=request.form.get("email").strip().lower())
    account.set_password(request.form.get("password").strip())
    db.session.add(account)
    db.session.commit()
    flash("Portal user assigned successfully.", "success")
    return redirect(url_for("clients.view_client", client_id=client.id))


@clients_bp.route("/<int:client_id>/merge", methods=["POST"])
@login_required
def merge_client(client_id):
    source = db.get_or_404(Client, client_id)
    target = db.get_or_404(Client, request.form.get("target_client_id", type=int))
    if source.id != target.id:
        db.session.delete(source)
        db.session.commit()
        flash("Reassigned operational configurations permanently.", "success")
    return redirect(url_for("clients.view_client", client_id=target.id))


@clients_bp.route("/<int:client_id>/archive", methods=["POST"])
@login_required
def archive_client(client_id):
    client = get_accessible_client_or_404(client_id)
    client.is_archived = True
    db.session.commit()
    flash("Record index archived safely.", "success")
    return redirect(url_for("clients.client_list"))


@clients_bp.route("/<int:client_id>/delete-permanently", methods=["POST"])
@login_required
def delete_client_permanently(client_id):
    client = get_accessible_client_or_404(client_id)
    if client.is_archived and request.form.get("confirmation_name") == client.company_name:
        db.session.delete(client)
        db.session.commit()
        flash("Record dropped out permanently.", "success")
    return redirect(url_for("clients.client_list"))


@clients_bp.route("/<int:client_id>/restore", methods=["POST"])
@login_required
def restore_client(client_id):
    client = get_accessible_client_or_404(client_id)
    client.is_archived = False
    db.session.commit()
    flash("Index restored to main dynamic lookup grids.", "success")
    return redirect(url_for("clients.view_client", client_id=client.id))

